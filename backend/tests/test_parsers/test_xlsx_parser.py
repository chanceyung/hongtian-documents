"""Unit tests for XLSX parser."""

from pathlib import Path

import pytest
import openpyxl
from openpyxl.styles import Font, PatternFill

from app.parsers.xlsx_parser import XLSXParser
from app.models.unified_document import UnifiedDocument


class TestXLSXParser:
    """Test suite for XLSX parser."""

    @pytest.fixture
    def sample_xlsx_with_content(self, tmp_path: Path) -> Path:
        """Create a sample XLSX file with data."""
        wb = openpyxl.Workbook()

        # Get active sheet
        ws = wb.active
        ws.title = "DataSheet"

        # Add headers with formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        ws['A1'] = "Name"
        ws['A1'].fill = header_fill
        ws['A1'].font = header_font

        ws['B1'] = "Age"
        ws['B1'].fill = header_fill
        ws['B1'].font = header_font

        ws['C1'] = "City"
        ws['C1'].fill = header_fill
        ws['C1'].font = header_font

        # Add data rows
        data_rows = [
            ["Alice", 28, "New York"],
            ["Bob", 32, "Los Angeles"],
            ["Charlie", 25, "Chicago"],
            ["Diana", 30, "Houston"],
        ]

        for row_data in data_rows:
            ws.append(row_data)

        # Save file
        file_path = tmp_path / "sample.xlsx"
        wb.save(file_path)
        wb.close()
        return file_path

    @pytest.fixture
    def empty_xlsx(self, tmp_path: Path) -> Path:
        """Create an empty XLSX file with blank sheet."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EmptySheet"

        file_path = tmp_path / "empty.xlsx"
        wb.save(file_path)
        wb.close()
        return file_path

    @pytest.fixture
    def xlsx_with_multiple_sheets(self, tmp_path: Path) -> Path:
        """Create XLSX with multiple sheets."""
        wb = openpyxl.Workbook()

        # First sheet
        ws1 = wb.active
        ws1.title = "FirstSheet"
        ws1.append(["Name", "Value"])
        ws1.append(["Item1", 100])
        ws1.append(["Item2", 200])

        # Second sheet
        ws2 = wb.create_sheet("SecondSheet")
        ws2.append(["Product", "Price"])
        ws2.append(["ProdA", 10.5])
        ws2.append(["ProdB", 20.8])

        # Third sheet (empty)
        ws3 = wb.create_sheet("EmptySheet")

        file_path = tmp_path / "multiple_sheets.xlsx"
        wb.save(file_path)
        wb.close()
        return file_path

    @pytest.mark.asyncio
    async def test_parser_xlsx_with_data_returns_unified_document(
        self,
        sample_xlsx_with_content: Path
    ) -> None:
        """Test parsing XLSX with data returns UnifiedDocument."""
        parser = XLSXParser()
        result: UnifiedDocument = await parser.parse(sample_xlsx_with_content)

        # Verify type
        assert isinstance(result, UnifiedDocument)

        # Verify tables extracted
        assert len(result.tables) > 0, "Should extract table from XLSX"

        # Verify fingerprint generated
        assert result.fingerprint != "", "Should generate fingerprint"

        # Verify parse method
        assert result.parse_method == "openpyxl", "Should use openpyxl parser"

    @pytest.mark.asyncio
    async def test_parser_xlsx_headers_correct(self, sample_xlsx_with_content: Path) -> None:
        """Test that table headers are correctly extracted."""
        parser = XLSXParser()
        result: UnifiedDocument = await parser.parse(sample_xlsx_with_content)

        assert len(result.tables) == 1, "Should extract one table"

        table = result.tables[0]
        assert len(table.headers) == 3, "Table should have 3 headers"

        header_texts = [h.content for h in table.headers]
        assert "Name" in header_texts, "Should extract 'Name' header"
        assert "Age" in header_texts, "Should extract 'Age' header"
        assert "City" in header_texts, "Should extract 'City' header"

    @pytest.mark.asyncio
    async def test_parser_xlsx_row_data_correct(self, sample_xlsx_with_content: Path) -> None:
        """Test that row data is correctly extracted."""
        parser = XLSXParser()
        result: UnifiedDocument = await parser.parse(sample_xlsx_with_content)

        table = result.tables[0]
        assert len(table.rows) == 4, "Table should have 4 data rows"

        # Check first row
        first_row = table.rows[0]
        first_row_content = [cell.content for cell in first_row.cells]
        assert "Alice" in first_row_content, "Should extract Alice"
        assert "28" in first_row_content, "Should extract age 28"
        assert "New York" in first_row_content, "Should extract New York"

        # Check last row
        last_row = table.rows[3]
        last_row_content = [cell.content for cell in last_row.cells]
        assert "Diana" in last_row_content, "Should extract Diana"
        assert "30" in last_row_content, "Should extract age 30"

    @pytest.mark.asyncio
    async def test_parser_xlsx_multiple_sheets_handling(self, xlsx_with_multiple_sheets: Path) -> None:
        """Test parsing XLSX with multiple sheets."""
        parser = XLSXParser()
        result: UnifiedDocument = await parser.parse(xlsx_with_multiple_sheets)

        # Should extract tables from non-empty sheets
        assert len(result.tables) == 2, "Should extract 2 tables (empty sheet skipped)"

        # Verify first table content
        first_table = result.tables[0]
        first_table_content = " ".join([cell.content for row in first_table.rows for cell in row.cells])
        assert "Item1" in first_table_content, "Should extract data from first sheet"

        # Verify second table content
        second_table = result.tables[1]
        second_table_content = " ".join([cell.content for row in second_table.rows for cell in row.cells])
        assert "ProdA" in second_table_content, "Should extract data from second sheet"

    @pytest.mark.asyncio
    async def test_parser_xlsx_empty_sheet_handling(self, xlsx_with_multiple_sheets: Path) -> None:
        """Test that empty sheets are handled correctly (skipped)."""
        parser = XLSXParser()
        result: UnifiedDocument = await parser.parse(xlsx_with_multiple_sheets)

        # Should only extract tables from non-empty sheets
        assert len(result.tables) == 2, "Should skip empty sheet"

        # Verify sheet names in metadata if available
        table_names = [t.metadata.get('sheet_name', '') if hasattr(t, 'metadata') and t.metadata else '' for t in result.tables]
        assert any("FirstSheet" in name or "SecondSheet" in name for name in table_names), "Should include sheet names"

    @pytest.mark.asyncio
    async def test_parser_xlsx_empty_document(self, empty_xlsx: Path) -> None:
        """Test parsing empty XLSX."""
        parser = XLSXParser()
        result: UnifiedDocument = await parser.parse(empty_xlsx)

        assert isinstance(result, UnifiedDocument)
        assert result.fingerprint != "", "Should still generate fingerprint for empty file"
        assert result.parse_method == "openpyxl"
        # Empty sheet should produce no tables
        assert len(result.tables) == 0, "Empty XLSX should have no tables"

    @pytest.mark.asyncio
    async def test_parser_xlsx_nonexistent_file(self, tmp_path: Path) -> None:
        """Test parsing non-existent XLSX file raises appropriate error."""
        parser = XLSXParser()
        nonexistent_file = tmp_path / "nonexistent.xlsx"

        with pytest.raises(FileNotFoundError):
            await parser.parse(nonexistent_file)

    @pytest.mark.asyncio
    async def test_parser_xlsx_numeric_values(self, tmp_path: Path) -> None:
        """Test that numeric values are correctly extracted."""
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(["Integer", "Float", "Formula"])
        ws.append([100, 3.14, "=A2+B2"])

        file_path = tmp_path / "numeric.xlsx"
        wb.save(file_path)
        wb.close()

        parser = XLSXParser()
        result: UnifiedDocument = await parser.parse(file_path)

        table = result.tables[0]
        first_row = table.rows[0]
        first_row_content = [cell.content for cell in first_row.cells]

        assert any("100" in content for content in first_row_content), "Should extract integer"
        assert any("3.14" in content for content in first_row_content), "Should extract float"

    @pytest.mark.asyncio
    async def test_parser_xlsx_large_table(self, tmp_path: Path) -> None:
        """Test parsing XLSX with large table."""
        wb = openpyxl.Workbook()
        ws = wb.active

        # Add header
        ws.append(["Column1", "Column2", "Column3"])

        # Add many rows
        for i in range(100):
            ws.append([f"Value1_{i}", f"Value2_{i}", f"Value3_{i}"])

        file_path = tmp_path / "large.xlsx"
        wb.save(file_path)
        wb.close()

        parser = XLSXParser()
        result: UnifiedDocument = await parser.parse(file_path)

        table = result.tables[0]
        assert len(table.rows) == 100, "Should extract all 100 rows"

        # Check first and last rows
        first_row_content = [cell.content for cell in table.rows[0].cells]
        assert "Value1_0" in first_row_content, "Should extract first row"

        last_row_content = [cell.content for cell in table.rows[99].cells]
        assert "Value1_99" in last_row_content, "Should extract last row"

    @pytest.mark.asyncio
    async def test_parser_xlsx_fingerprint_uniqueness(self, sample_xlsx_with_content: Path) -> None:
        """Test that fingerprint is unique and consistent for same file."""
        parser = XLSXParser()

        # Parse same file twice
        result1: UnifiedDocument = await parser.parse(sample_xlsx_with_content)
        result2: UnifiedDocument = await parser.parse(sample_xlsx_with_content)

        # Fingerprints should be identical for same file
        assert result1.fingerprint == result2.fingerprint, "Fingerprint should be consistent"

        # Fingerprint should not be empty
        assert len(result1.fingerprint) > 0, "Fingerprint should not be empty"

    @pytest.mark.asyncio
    async def test_parser_xlsx_merged_cells(self, tmp_path: Path) -> None:
        """Test handling of merged cells."""
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(["Title", "", "Subtitle"])
        ws.append(["Data1", "Data2", "Data3"])

        # Merge cells in first row
        ws.merge_cells('B1:C1')

        file_path = tmp_path / "merged.xlsx"
        wb.save(file_path)
        wb.close()

        parser = XLSXParser()
        result: UnifiedDocument = await parser.parse(file_path)

        # Should still extract data
        assert len(result.tables) > 0, "Should extract table even with merged cells"

        table = result.tables[0]
        assert len(table.headers) > 0, "Should extract headers"